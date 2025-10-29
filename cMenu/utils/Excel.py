from typing import (Dict, List, Any, )

from PySide6.QtCore import (
    QAbstractTableModel, 
    )

from openpyxl import (Workbook, )
from openpyxl.styles import PatternFill, Font, fills, colors
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH

from .cQModels import (SQLAlchemyTableModel, )

ExcelWorkbook_fileext = ".XLSX"


def Excelfile_fromqs(qset:SQLAlchemyTableModel|List[Dict[str, Any]], flName:str|None = None,
                     freezecols:int = 0, returnFileName: bool = False) -> Workbook|str:
    """
    qset: a QAbstractTableModel or list of dictionaries
    flName: the name of the file to be built (WITHOUT extension!).  It's stored on the server.  If it's to be dl'd, the caller does that
    freezecols = 0: the number of columns to freeze to the left
    The top row contains the field names, is always frozen, is bold and is shaded grey

    used to Return the name of the Workbook file (with extension).  Once I start building in errorchecking and exceptions, other returns may be possible
    Returns the Workbook file (with extension)
    """

    # far easier to process a list of dictionaries, so...
    if isinstance(qset,QAbstractTableModel):
        # make this a util
        qlist = qset.getDataAsList()
    elif isinstance(qset,list):
        qlist = qset
    else:
        return None
    if qlist:
        if not isinstance(qlist[0],dict):
            # review this later ...
            try:
                qlist = [n.__dict__ for n in qlist]
            except:
                qlist = []

    # create empty workbook with an empty worksheet
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    # header row is names of columns
    if qlist:
        fields = list(qlist[0])
        ws.append(fields)

        # append each row
        for row in qlist:
            ws.append(list(row.values())) # type: ignore

        # make header row bold, shade it grey, freeze it
        # ws.show_gridlines = True  #Nope - this is a R/O attribute
        for cell in ws[1]: # type: ignore
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type=fills.FILL_SOLID,
                            start_color=colors.Color("00808080"),
                            end_color=colors.Color("00808080")
                            )
        #TODO: convert row1 and cols:freezecols to an address (A=0, B=1, C=2 etc) for line below
        ws.freeze_panes ='A2'
        #TODO: if freezecols passed, freeze them, too


    # save the workbook
    if flName:
        wb.save(flName+ExcelWorkbook_fileext)
    
    if returnFileName:
        # close the workbook
        wb.close()
        # and return file Name to the caller
        return (flName+ExcelWorkbook_fileext if flName else "")
    else:
        # return the workbook itself
        return wb
    #endif returnFileName


class UpldSprdsheet():
    """Base class for handling spreadsheet uploads with field validation.
    
    This class provides functionality for processing uploaded spreadsheets,
    validating field types, and cleaning data according to defined rules.
    
    Attributes:
        TargetModel: The target ORM model class for the spreadsheet data.
        SprdsheetDateEpoch: Date epoch used for spreadsheet date conversion.
        SprdsheetFlds (dict): Dictionary mapping spreadsheet field names to field descriptors.
    """
    TargetModel = None
    SprdsheetDateEpoch = WINDOWS_EPOCH

    def SprdsheetFldDescriptor_creator(self, ModelFldName, AllowedTypes):
        """Create a field descriptor for spreadsheet field validation.
        
        Args:
            ModelFldName (str): The name of the field in the TargetModel.
            AllowedTypes: List of tuples (type, cleanproc) specifying allowed types
                and their cleaning procedures. Empty list if any string is allowed.
        
        Returns:
            dict: Field descriptor dictionary with ModelFldName and AllowedTypes.
        """
        return  {
            # 'SprdsheetName': None,    # nope, this will be the index of SprdsheetFlds
            'ModelFldName': ModelFldName,
            'AllowedTypes': AllowedTypes,     
        }
    
    SprdsheetFlds = {}  # key will be the SprdsheetName, value is a SprdsheetFldDescriptor

    def cleanupfld(self, fld, val):
        """Clean and validate a field value according to its allowed types.
        
        Args:
            fld: Field name to clean.
            val: Value to clean and validate.
        
        Returns:
            tuple: (usefld, cleanval) where usefld is True if the field should be used,
                and cleanval is the cleaned value.
        """
        usefld = False
        cleanval = None
        
        if fld not in self.SprdsheetFlds:
            # just feed the value back
            usefld = True
            cleanval = val
        elif not self.SprdsheetFlds[fld]['AllowedTypes']:
            usefld = (val is not None)
            if usefld: cleanval = str(val)
        else:
            for type, cleanproc in self.SprdsheetFlds[fld]['AllowedTypes']:
                if isinstance(val, type):
                    usefld = True
                    cleanval = cleanproc(val)
                    break
                #endif instance(val, type)
            #endfor type, cleanproc
        #endif fld not in self.SprdsheetFlds

    def process_spreadsheet(self, SprsheetName):
        """Process a spreadsheet file.
        
        Args:
            SprsheetName: Name or path of the spreadsheet to process.
        
        Note:
            This method is not yet implemented.
        """
        pass

